import openpyxl
import os
from datetime import datetime
from workflow_tracker import log_info

BLACKLIST_FILE = "blacklist.xlsx"

# In-memory cache — loaded once at startup, updated on each new entry
_blacklist_url_cache = None
_cache_loaded = False

def initialize_blacklist():
    global _blacklist_url_cache, _cache_loaded
    if not os.path.exists(BLACKLIST_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Blacklist"
        ws.append(["Job Title", "Company", "Portal", "URL", "Reason", "Date Added"])
        wb.save(BLACKLIST_FILE)
        log_info("[Blacklist] Created new blacklist.xlsx")
        _blacklist_url_cache = set()
        _cache_loaded = True
    else:
        if not _cache_loaded:
            _blacklist_url_cache = load_existing_blacklist_urls()
            _cache_loaded = True
            log_info("[Blacklist] blacklist.xlsx loaded into cache — " + str(len(_blacklist_url_cache)) + " entries")

def load_existing_blacklist_urls():
    urls = set()
    if not os.path.exists(BLACKLIST_FILE):
        return urls
    wb = openpyxl.load_workbook(BLACKLIST_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3]:
            urls.add(row[3].strip())
    return urls

def add_to_blacklist(job, reason="Blacklisted by keyword"):
    global _blacklist_url_cache
    initialize_blacklist()
    url = job.get("url", "").strip()
    if url and url in _blacklist_url_cache:
        return
    wb = openpyxl.load_workbook(BLACKLIST_FILE)
    ws = wb.active
    ws.append([
        job.get("title", ""),
        job.get("company", ""),
        job.get("portal", ""),
        job.get("url", ""),
        reason,
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(BLACKLIST_FILE)
    if url:
        _blacklist_url_cache.add(url)