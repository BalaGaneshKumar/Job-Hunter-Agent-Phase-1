"""
dashboard_backup.py — "One Drive Uploader" backup
Reads jobs.csv + blacklist.csv and writes a single, styled .xlsx with one
tab per status into the local OneDrive-synced folder, so OneDrive picks it
up and syncs it to the cloud automatically.
"""

import os
import csv
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dashboard_workflows import JOBS_CSV, BLACKLIST_CSV
from career_tracker import load_career_apps, load_career_sites

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BACKUP_DIR  = r"C:\Users\balag\OneDrive\Applied"
BACKUP_FILE = os.path.join(BACKUP_DIR, "Applied_Jobs_Backup.xlsx")

# status -> (sheet name, tab color hex, header fill hex)
STATUS_TABS = [
    ('Scraped',    'Scraped',    '1565C0'),
    ('Eligible',   'Eligible',   '2E7D32'),
    ('Applied',    'Applied',    '6A1B9A'),
    ('Ineligible', 'Ineligible', 'C62828'),
    ('Hold',       'Hold',       'F57F17'),
]
BLACKLIST_TAB = ('Blacklist', 'C62828')

HEADER_BG = 'FFFFFF'  # overridden per-tab below via header fill hex
BAND_BG   = 'F7F9FC'
THIN  = Side(border_style='thin', color='90A4AE')
THICK = Side(border_style='thick', color='2C3E50')


def _apply_grid_border(ws, n_rows, n_cols):
    """Thick border around the outside of the table, thin gridlines inside."""
    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = Border(
                top=THICK if r == 1 else THIN,
                bottom=THICK if r == n_rows else THIN,
                left=THICK if c == 1 else THIN,
                right=THICK if c == n_cols else THIN,
            )


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_sheet(ws, headers, rows, urls, hyperlink_col, tab_color, header_color):
    """rows: list of value-lists matching `headers` length.
    urls: parallel list giving the link target for `hyperlink_col` in each row."""
    ws.sheet_properties.tabColor = tab_color

    # Header row
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=header_color)
        c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 20

    # Data rows
    for r_i, (row, url) in enumerate(zip(rows, urls), start=2):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            if r_i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=BAND_BG)
            if c_i == hyperlink_col and val and url:
                cell.hyperlink = url
                cell.font = Font(color='2196F3', underline='single')

    _apply_grid_border(ws, len(rows) + 1, len(headers))

    if rows:
        ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(headers)), len(rows) + 1)
    ws.freeze_panes = 'A2'


def _read_csv(path):
    if not path or not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8', newline='') as f:
        return list(csv.DictReader(f))


def run_onedrive_backup():
    """Builds Applied_Jobs_Backup.xlsx (one tab per status) and saves it
    into the local OneDrive folder. Returns {'ok', 'path'/'error', 'counts'}."""
    print('[Backup] One Drive Uploader triggered...', flush=True)
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)

        jobs = _read_csv(JOBS_CSV)
        bl_rows = _read_csv(BLACKLIST_CSV)

        wb = Workbook()
        wb.remove(wb.active)
        counts = {}

        base_headers = ['Company', 'Job Title', 'Portal', 'Applied Type', 'Location', 'Last Checked']
        widths = [22, 40, 12, 13, 20, 16]
        for status, sheet_name, header_color in STATUS_TABS:
            matched = [j for j in jobs if (j.get('Status') or 'Scraped') == status]
            rows, urls = [], []
            for j in matched:
                ie = j.get('Internal/External', '')
                ie = ie if ie in ('Internal', 'External') else ''
                rows.append([j.get('Company', ''), j.get('Job Title', ''),
                             j.get('Portal', ''), ie, j.get('Location', ''),
                             j.get('Last Checked', '')])
                urls.append(j.get('URL', ''))
            ws = wb.create_sheet(sheet_name)
            _style_sheet(ws, base_headers, rows, urls, hyperlink_col=2,
                         tab_color=header_color, header_color=header_color)
            _autosize(ws, widths)
            counts[status] = len(rows)

        # Blacklist: Company | Job Title (hyperlink) | Portal — from blacklist.csv
        # blacklist.csv header order: Job Title, Company, Portal, URL
        bl_data, bl_urls = [], []
        for j in bl_rows:
            cols = list(j.keys())
            title   = j.get(cols[0], '') if len(cols) > 0 else ''
            company = j.get(cols[1], '') if len(cols) > 1 else ''
            portal  = j.get(cols[2], '') if len(cols) > 2 else ''
            url     = j.get(cols[3], '') if len(cols) > 3 else ''
            bl_data.append([company, title, portal])
            bl_urls.append(url)
        ws = wb.create_sheet(BLACKLIST_TAB[0])
        _style_sheet(ws, ['Company', 'Job Title', 'Portal'], bl_data, bl_urls,
                     hyperlink_col=2, tab_color=BLACKLIST_TAB[1],
                     header_color=BLACKLIST_TAB[1])
        _autosize(ws, [22, 40, 12])
        counts['Blacklist'] = len(bl_data)

        # Applications tab — from career_apps.csv (via career_tracker)
        apps = load_career_apps()
        app_rows = [[a.get('company', ''),
                     (a.get('credential', '') + ' ' + a.get('credtype', '')).strip(),
                     a.get('role', ''), a.get('location', ''), a.get('date', ''), a.get('stage', '')]
                    for a in apps]
        app_urls = [a.get('link', '') for a in apps]
        ws = wb.create_sheet('Applications')
        _style_sheet(ws, ['Company', 'Credential', 'Role', 'Location', 'Applied', 'Stage'],
                     app_rows, app_urls, hyperlink_col=1,
                     tab_color='3498DB', header_color='3498DB')
        _autosize(ws, [22, 16, 26, 18, 14, 14])
        counts['Applications'] = len(app_rows)

        # Job Sites tab — from career_sites.csv (via career_tracker)
        sites = load_career_sites()
        site_rows = [[s.get('name', ''), s.get('cred', ''), s.get('notes', '')] for s in sites]
        site_urls = [s.get('url', '') for s in sites]
        ws = wb.create_sheet('Job Sites')
        _style_sheet(ws, ['Site', 'Credential Hint', 'Notes'],
                     site_rows, site_urls, hyperlink_col=1,
                     tab_color='16A085', header_color='16A085')
        _autosize(ws, [24, 20, 40])
        counts['Job Sites'] = len(site_rows)

        # Write outside the synced folder first — if the temp file lived
        # inside BACKUP_DIR, OneDrive could pick it up and start syncing it
        # as its own file a moment before the final rename lands, and that
        # overlap is what triggers "-<name>" conflict copies even when each
        # individual sync finishes in seconds. Staging elsewhere means
        # OneDrive only ever observes one clean write to the real filename.
        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(tmp_path)
        try:
            os.replace(tmp_path, BACKUP_FILE)
        except PermissionError:
            os.remove(tmp_path)
            print('[Backup] FAILED — file is open/locked: ' + BACKUP_FILE, flush=True)
            return {'ok': False, 'error':
                    'Applied_Jobs_Backup.xlsx is open in Excel (or locked) — close it and try again.'}
        except OSError as e:
            # os.replace across different drives isn't atomic — fall back
            # to a same-folder temp file (old behavior) if system temp is
            # on a different volume than the OneDrive folder.
            os.remove(tmp_path)
            fallback_tmp = BACKUP_FILE + '.tmp'
            wb.save(fallback_tmp)
            os.replace(fallback_tmp, BACKUP_FILE)
        print('[Backup] Saved ' + BACKUP_FILE + ' — ' + str(counts), flush=True)
        print('[Backup] Completed.', flush=True)
        return {'ok': True, 'path': BACKUP_FILE, 'counts': counts}
    except Exception as e:
        print('[Backup] ERROR: ' + str(e), flush=True)
        return {'ok': False, 'error': str(e)}
