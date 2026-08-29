import os
import csv
import openpyxl
from datetime import datetime
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from workflow_tracker import log_info

SCRAPER_FILE   = 'scraper_jobs.xlsx'
BLACKLIST_FILE = 'blacklist.xlsx'
JOBS_CSV       = 'jobs.csv'

COLUMNS = [
    'Job ID', 'Job Title', 'Company', 'Portal', 'Location',
    'URL', 'Status', 'Internal/External', 'Applied Date', 'Last Checked',
    'Description', 'Notes'
]

# ---------------------------------------------------------------------------
# Scraper Excel helpers
# ---------------------------------------------------------------------------
def initialize_scraper_excel():
    if not os.path.exists(SCRAPER_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Jobs'
        ws.append(COLUMNS)
        wb.save(SCRAPER_FILE)
        log_info('[Tracker] Created new scraper_jobs.xlsx')
    else:
        log_info('[Tracker] scraper_jobs.xlsx already exists')


def load_existing_jobs():
    """Returns (urls_set, company_title_pairs_set) for deduplication."""
    urls  = set()
    pairs = set()
    if not os.path.exists(SCRAPER_FILE):
        return urls, pairs
    wb = openpyxl.load_workbook(SCRAPER_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5]:
            urls.add(row[5])
        company = str(row[2] or '').strip().lower()
        title   = str(row[1] or '').strip().lower()
        if company and title and company != 'unknown':
            pairs.add((company, title))
    return urls, pairs


def batch_write_jobs(jobs, filename=SCRAPER_FILE):
    """Write multiple jobs in one open/save cycle."""
    if not jobs:
        return 0
    try:
        wb  = openpyxl.load_workbook(filename)
        ws  = wb.active
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        for job in jobs:
            ws.append([
                sanitize(job.get('job_id',      '')),
                sanitize(job.get('title',        '')),
                sanitize(job.get('company',      '')),
                sanitize(job.get('portal',       '')),
                sanitize(job.get('location',     '')),
                sanitize(job.get('url',          '')),
                sanitize(job.get('status',       'Scraped')),
                sanitize(job.get('applied_date', '')),
                now,
                sanitize(job.get('description', '')),
                sanitize(job.get('notes',        ''))
            ])
        wb.save(filename)
        return len(jobs)
    except Exception as e:
        log_info('[Tracker] Batch write error: ' + str(e)[:50])
        return 0


# ---------------------------------------------------------------------------
# Sync scraper_jobs.xlsx → jobs.csv  (no more jobs.xlsx step)
# ---------------------------------------------------------------------------
def sync_to_master():
    if not os.path.exists(SCRAPER_FILE):
        log_info('[Tracker] scraper_jobs.xlsx not found, skipping sync')
        return 0

    try:
        src    = openpyxl.load_workbook(SCRAPER_FILE)
        src_ws = src.active
    except Exception as e:
        log_info('[Tracker] Sync read error: ' + str(e)[:50])
        return 0

    # Load existing jobs.csv into memory
    existing = {}   # url -> row dict
    if os.path.exists(JOBS_CSV):
        try:
            with open(JOBS_CSV, 'r', encoding='utf-8', newline='') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    url = row.get('URL', '').strip()
                    if url:
                        existing[url] = row
        except Exception as e:
            log_info('[Tracker] CSV read error: ' + str(e)[:50])

    added   = 0
    patched = 0
    now     = datetime.now().strftime('%Y-%m-%d %H:%M')
    synced_urls = set()

    for src_row in src_ws.iter_rows(min_row=2, values_only=True):
        url = src_row[5]
        if not url or url in synced_urls:
            continue
        synced_urls.add(url)

        src_company     = str(src_row[2] or '').strip()
        src_description = str(src_row[9] or '').strip()

        if url not in existing:
            existing[url] = {
                'Job ID'            : sanitize(src_row[0]) or '',
                'Job Title'         : sanitize(src_row[1]) or '',
                'Company'           : sanitize(src_row[2]) or '',
                'Portal'            : sanitize(src_row[3]) or '',
                'Location'          : sanitize(src_row[4]) or '',
                'URL'               : sanitize(src_row[5]) or '',
                'Status'            : sanitize(src_row[6]) or 'Scraped',
                'Internal/External' : '',   # filled later by the Validator, never by sync
                'Applied Date'      : sanitize(src_row[7]) or '',
                'Last Checked'      : now,
                'Description'       : sanitize(src_row[9]) or '',
                'Notes'             : sanitize(src_row[10]) or ''
            }
            added += 1
        else:
            row     = existing[url]
            updated = False

            company_missing = not row.get('Company','').strip() or row.get('Company','').strip().lower() == 'unknown'
            if company_missing and src_company and src_company.lower() != 'unknown':
                row['Company'] = sanitize(src_company)
                updated = True

            desc_missing = not row.get('Description','').strip()
            if desc_missing and src_description:
                row['Description'] = sanitize(src_description)
                updated = True

            if updated:
                row['Last Checked'] = now
                patched += 1

    # Write merged data back to jobs.csv
    try:
        with open(JOBS_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS)
            writer.writeheader()
            writer.writerows(existing.values())
        log_info('[Tracker] Synced ' + str(added) + ' new, patched ' + str(patched) + ' existing → jobs.csv')
    except Exception as e:
        log_info('[Tracker] CSV write error: ' + str(e)[:50])

    # Export blacklist.xlsx → blacklist.csv
    export_blacklist_csv()
    return added


# ---------------------------------------------------------------------------
# Blacklist CSV export
# ---------------------------------------------------------------------------
def export_blacklist_csv():
    if not os.path.exists(BLACKLIST_FILE):
        return
    try:
        wb  = openpyxl.load_workbook(BLACKLIST_FILE)
        ws  = wb.active
        csv_file = BLACKLIST_FILE.replace('.xlsx', '.csv')
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([v if v is not None else '' for v in row])
        log_info('[Tracker] Exported to blacklist.csv')
    except Exception as e:
        log_info('[Tracker] Blacklist CSV export error: ' + str(e)[:50])


# ---------------------------------------------------------------------------
# Sanitize
# ---------------------------------------------------------------------------
def sanitize(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


# ---------------------------------------------------------------------------
# Entry point — manual sync
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    sync_to_master()
